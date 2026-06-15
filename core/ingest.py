import os
import json
import re
import openpyxl
from langchain_core.documents import Document
from langchain_google_genai import ChatGoogleGenerativeAI
from dotenv import load_dotenv

load_dotenv()

api_key = os.getenv("GEMINI_API_KEY")


# Prompt: Directions for LLM to parse worksheet

_EXTRACTION_SYSTEM_PROMPT = """
You are a fitness data extraction assistant.
You will be given a raw text dump of a spreadsheet sheet from a workout program.
The spreadsheet may be formatted in any way — days as columns, days as row blocks,
exercises stacked vertically, Weight/Reps on separate rows, merged cells shown as
repeated values, sparse grids, free-text notes, etc.
 
Your job is to extract every workout session on the sheet and return ONLY a JSON array.
Do not include any explanation, markdown, or code fences — just the raw JSON array.
 
Each element in the array must be an object with these fields:
{
  "week": <string, e.g. "Week 1" or null if unknown>,
  "day": <string, e.g. "Day 1" or "Monday" or the day label from the sheet>,
  "session_name": <string describing the session, e.g. "Legs - Quad Focus" or null>,
  "exercises": [
    {
      "name": <string, exercise name>,
      "prescribed_sets_reps": <string, e.g. "3 x 6-8" or null if not shown>,
      "actual_weight": <string, exactly as written, e.g. "65,75,75" or "70 assisted" or null>,
      "actual_reps": <string, exactly as written, e.g. "3x10" or "2x6, 1x2" or null>,
      "notes": <string, any extra notes on form, intensity, or null>
    }
  ],
  "is_rest_day": <boolean, true if this is a rest day with no exercises>
}
 
Rules:
- Include rest days as entries with is_rest_day=true and an empty exercises list.
- If weight or reps were not logged (cell is blank), use null — do not invent values.
- Preserve free-text weight values exactly as written (e.g. "Pendulum 10", "7 on assisted").
- If you cannot determine what week this sheet belongs to, infer it from the sheet name provided.
- Consolidate all exercises for a given day into one session object.
- Return an empty array [] only if the sheet contains no workout data at all.
"""

_EXTRACTION_USER_TEMPLATE = """ Sheet name: {sheet_name}
 
Raw cell contents (row number | col number | value):
{cell_dump}
 
Extract all workout sessions from this sheet and return the JSON array."""


def load_workout_documents(xlsx_path: str, api_key: str) -> list[Document]:
    """
    Load xlsx file and return LangChain Documents. Each doc represents one workout session.
    """

    llm = ChatGoogleGenerativeAI(
        model="gemini-2.5-flash",
        google_api_key=api_key,
        temperature=0
    )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_docs = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"[Ingest] Processing Sheet: '{sheet_name}'...")

        cell_dump = _dump_sheet(ws)
        if not cell_dump.strip():
            print(f"[Ingest]    Sheet '{sheet_name}' appears empty, skipping.")
            continue

        sessions = _extract_sessions_via_llm(llm, sheet_name, cell_dump)
        if not sessions:
            print(f"[Ingest]    No sessions found in '{sheet_name}'.")
            continue

        docs = _sessions_to_documents(sessions, sheet_name, xlsx_path)
        print(f"[Ingest]    Extracted {len(docs)} sessions from '{sheet_name}'.")
        all_docs.extend(docs)
    print(f"[Ingest] Total documents built: {len(all_docs)}")
    return all_docs

        




# Helper Functions

#Converts spreadsheet into a raw text dump
def _dump_sheet(ws) -> str:
    """
    Converts a worksheet into a text representation for the LLM.
    Format: "row R | col C | <value>" One line per non-empty cell
    """

    lines = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                lines.append(f"row {cell.row} | col {cell.column} | {cell.value}")
    return "\n".join(lines)


# LLM Extraction

def _extract_sessions_via_llm(llm, sheet_name: str, cell_dump: str) -> list[dict]:
    """
    Send raw cell dump to LLM and get back a list of session dicts.
    """

    prompt = _EXTRACTION_USER_TEMPLATE.format(
        sheet_name=sheet_name,
        cell_dump=cell_dump
    )

    messages = [
        ("system", _EXTRACTION_SYSTEM_PROMPT),
        ("human", prompt)
    ]

    response = llm.invoke(messages)
    raw_text = response.content.strip()

    return _parse_json_response(raw_text, sheet_name)


def _parse_json_response(raw_text: str, sheet_name: str) -> list[dict]:
    """
    Parses LLM's JSON response.
    """
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw_text, flags=re.MULTILINE)
    cleaned = re.sub(r"\s*```$", "", cleaned, flags=re.MULTILINE).strip()

    try:
        res = json.loads(cleaned)
        if isinstance(res, list):
            return res
        if isinstance(res, dict):
            for key in ("sessions", "data", "workouts"):
                if key in res and isinstance(res[key], list):
                    return res[key]
    except json.JSONDecodeError as e:
        print(f"[Ingest]    WARNING: JSON parse failed for sheet '{sheet_name}': {e}")
        print(f"[Ingest]    Raw response snippet: {raw_text[:300]}")
    
    return []

#Converts sessions to langchain docs
def _sessions_to_documents(sessions: list[dict], sheet_name: str, source_path: str) -> list[Document]:
    """
    Convert extracted session dicts into LangChain Documents.
    page_content is human-readable narrative
    metadata carries structued fields for filtering
    """

    documents = []

    for session in sessions:
        if not isinstance(session, dict):
            continue
        is_rest = session.get("is_rest_day", False)
        week = session.get("week") or sheet_name
        day = session.get("day") or "Unknown day"
        session_name = session.get("session_name") or "General Workout Session"
        exercises = session.get("exercises") or []

        all_exercise_names = [ex.get("name", "") for ex in exercises if ex.get("name")]

        if is_rest:
            content = [
                f"Week: {week}",
                f"Day: {day}",
                f"Workout: {session_name}",
                "Exercise: Rest Day",
                "Prescription: N/A",
                "Performed: Rest, recovery, and mobile conditioning.",
                "Other exercises this day: None"
            ]
            content = "\n".join(content)

            metadata = {
                "week": week,
            "day": day,
            "session_name": session_name,
            "is_rest_day": True,
            "exercise_name": "Rest Day",
            "exercise_names": ["Rest Day"], # Keeps pipeline.py compatibility
            "sheet_name": sheet_name,
            "source": source_path,
            }
            documents.append(Document(page_content=content, metadata=metadata))
        else:
            for ex in exercises:
                current_name = ex.get("name", "Unknown exercise").strip()
                prescribed = ex.get("prescribed_sets_reps") or "Not Specified"
                weight = ex.get("actual_weight") or ""
                reps = ex.get("actual_reps") or ""
                notes = ex.get("notes") or ""

                other_exercises = [name for name in all_exercise_names if name != current_name]
                other_ex_str = ", ".join(other_exercises) if other_exercises else "None"
                
                performed_string = "No log data provided."
                if weight or reps:
                    weight_list = [w.strip() for w in str(weight).split(",") if w.strip()]
                    reps_list = [r.strip() for r in str(reps).split(",") if r.strip()]

                    if len(weight_list) == len(reps_list) and len(weight_list) > 0:
                        set_strings = [f"{w} lbs x {r}" for w, r in zip(weight_list, reps_list)]
                        performed_string = ", ".join(set_strings)
                    else:
                        perf_parts = []
                        if weight_list:
                            perf_parts.append(f"Weight: {weight_list[0]}")
                        if reps_list:
                            perf_parts.append(f"Reps: {reps_list[0]}")
                        performed_string = "| ".join(perf_parts)
                if notes:
                    performed_string += f" | Notes: {notes}"
                
                content = [
                    f"Week: {week}",
                    f"Day: {day}",
                    f"Workout: {session_name}",
                    f"Exercise: {current_name}",
                    f"Prescription: {prescribed}",
                    f"Performed: {performed_string}",
                    f"Other exercises this day: {other_ex_str}"
                ]
                content = "\n".join(content)
        
                metadata = {
                    "week":           week,
                    "day":            day,
                    "session_name":   session_name,
                    "is_rest_day":    False,
                    "exercise_name": current_name,
                    "exercise_names": all_exercise_names, # For easier retrieval filtering
                    "sheet_name":     sheet_name,
                    "source":         source_path,
                }

                documents.append(Document(page_content=content, metadata=metadata))

    return documents

